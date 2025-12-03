import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3


def center_window(win, width=900, height=650):
    win.update_idletasks()
    screen_width = win.winfo_screenwidth()
    screen_height = win.winfo_screenheight()

    x = (screen_width // 2) - (width // 2)
    y = (screen_height // 2) - (height // 2)

    win.geometry(f"{width}x{height}+{x}+{y}")


def connect_db():
    conn = sqlite3.connect('quan_ly_benh_nhan.db')
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS benh_nhan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ma_bn TEXT,
            ho_ten TEXT,
            gioi_tinh TEXT,
            ngay_sinh TEXT,
            so_dt TEXT,
            dia_chi TEXT,
            chan_doan TEXT,
            ngay_nhap_vien TEXT,
            ghi_chu TEXT
        )
    """)
    conn.commit()
    return conn, cursor

class PatientManagerApp:
    def __init__(self, master):
        self.master = master
        master.title("Quản lý bệnh nhân")
        master.geometry("900x650")

        self.conn, self.cursor = connect_db()
        self.entries = {}

        self.input_frame = tk.Frame(master, padx=10, pady=10)
        self.input_frame.pack(fill='x', padx=10, pady=10)

        self._create_input_fields()
        self._create_action_buttons()
        self._create_table_view()

        self.load_data()

    def _create_input_fields(self):
        fields = [
            ("Mã BN", "ma_bn", 0, 0, 0, 1),
            ("Họ & tên", "ho_ten", 1, 0, 1, 1),
            ("Giới tính", "gioi_tinh", 2, 0, 2, 1),
            ("Ngày sinh (YYYY-MM-DD)", "ngay_sinh", 3, 0, 3, 1),
            ("Số ĐT", "so_dt", 4, 0, 4, 1),

            ("Địa chỉ", "dia_chi", 0, 2, 0, 3),
            ("Chẩn đoán", "chan_doan", 1, 2, 1, 3),
            ("Ngày nhập viện (YYYY-MM-DD)", "ngay_nhap_vien", 2, 2, 2, 3),
            ("Ghi chú", "ghi_chu", 3, 2, 3, 3),
        ]

        for label_text, name, r, c, er, ec in fields:
            tk.Label(self.input_frame, text=label_text).grid(row=r, column=c, sticky='w', padx=5, pady=5)

            if name == "gioi_tinh":
                combo = ttk.Combobox(self.input_frame, values=["Nam", "Nữ", "Khác"], width=30, state="readonly")
                combo.grid(row=er, column=ec, sticky='w', padx=5, pady=5)
                combo.set("Nam")
                self.entries[name] = combo
            else:
                entry = tk.Entry(self.input_frame, width=33)
                entry.grid(row=er, column=ec, sticky='w', padx=5, pady=5)
                self.entries[name] = entry

        self.input_frame.grid_columnconfigure(1, weight=1)
        self.input_frame.grid_columnconfigure(3, weight=1)

    def _create_action_buttons(self):
        row_btn = 5

        ttk.Button(self.input_frame, text="Cập nhật", command=self.update_patient).grid(row=row_btn, column=1, sticky='w')

        ttk.Button(self.input_frame, text="Xóa", command=self.delete_patient).grid(row=row_btn, column=1, padx=80, sticky='w')
        ttk.Button(self.input_frame, text="Làm mới", command=self.clear_fields).grid(row=row_btn, column=1, padx=140, sticky='w')

        ttk.Button(self.input_frame, text="Xuất Excel", command=self.export_excel).grid(row=row_btn, column=0, padx=90, sticky='w')

        tk.Label(self.input_frame, text="Tìm kiếm:").grid(row=row_btn, column=2, sticky='e')
        self.search_entry = tk.Entry(self.input_frame, width=18)
        self.search_entry.grid(row=row_btn, column=3, sticky='w')
        ttk.Button(self.input_frame, text="Tìm", command=self.search_data).grid(row=row_btn, column=3, padx=90, sticky='w')
        ttk.Button(self.input_frame, text="Tải tất cả", command=self.load_data).grid(row=row_btn, column=3, padx=150, sticky='w')
        ttk.Button(self.input_frame, text="Thêm", command=self.add_patient)\
        .grid(row=row_btn, column=0, pady=10, padx=5, sticky='w')

    def _create_table_view(self):
        self.table_frame = tk.Frame(self.master)
        self.table_frame.pack(fill='both', expand=True, padx=10, pady=10)

        columns = ("id", "ma_bn", "ho_ten", "gioi_tinh", "ngay_sinh", "so_dt",
                   "dia_chi", "chan_doan", "ngay_nhap_vien", "ghi_chu")

        self.tree = ttk.Treeview(self.table_frame, columns=columns, show="headings")

        for col in columns:
            self.tree.heading(col, text=col.upper())

        for col in columns:
            self.tree.column(col, anchor='center', width=120)

        vsb = ttk.Scrollbar(self.table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(self.table_frame, orient="horizontal", command=self.tree.xview)

        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.pack(fill='both', expand=True, side='left')
        vsb.pack(side='right', fill='y')
        hsb.pack(side='bottom', fill='x')

        self.tree.bind("<<TreeviewSelect>>", self.on_select_item)

    def load_data(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        self.cursor.execute("SELECT * FROM benh_nhan")
        for row in self.cursor.fetchall():
            self.tree.insert("", tk.END, values=row)

    def add_patient(self):
        data = [
            self.entries["ma_bn"].get(),
            self.entries["ho_ten"].get(),
            self.entries["gioi_tinh"].get(),
            self.entries["ngay_sinh"].get(),
            self.entries["so_dt"].get(),
            self.entries["dia_chi"].get(),
            self.entries["chan_doan"].get(),
            self.entries["ngay_nhap_vien"].get(),
            self.entries["ghi_chu"].get()
        ]

        if not data[0] or not data[1]:
            messagebox.showerror("Lỗi", "Mã BN và Họ tên không được để trống!")
            return

        try:
            self.cursor.execute("""
                INSERT INTO benh_nhan
                (ma_bn, ho_ten, gioi_tinh, ngay_sinh, so_dt, dia_chi, chan_doan, ngay_nhap_vien, ghi_chu)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, data)
            self.conn.commit()
            self.load_data()
            self.clear_fields()
            messagebox.showinfo("Thành công", "Đã thêm bệnh nhân.")
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))

    def update_patient(self):
        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showerror("Lỗi", "Vui lòng chọn dòng cần cập nhật.")
            return

        record_id = self.tree.item(selected_item)["values"][0]

        data = [
            self.entries["ma_bn"].get(),
            self.entries["ho_ten"].get(),
            self.entries["gioi_tinh"].get(),
            self.entries["ngay_sinh"].get(),
            self.entries["so_dt"].get(),
            self.entries["dia_chi"].get(),
            self.entries["chan_doan"].get(),
            self.entries["ngay_nhap_vien"].get(),
            self.entries["ghi_chu"].get(),
            record_id
        ]

        try:
            self.cursor.execute("""
                UPDATE benh_nhan 
                SET ma_bn=?, ho_ten=?, gioi_tinh=?, ngay_sinh=?, so_dt=?, 
                    dia_chi=?, chan_doan=?, ngay_nhap_vien=?, ghi_chu=?
                WHERE id=?
            """, data)
            self.conn.commit()
            self.load_data()
            self.clear_fields()
            messagebox.showinfo("Thành công", "Đã cập nhật.")
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))

    def delete_patient(self):
        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showerror("Lỗi", "Vui lòng chọn dòng để xóa.")
            return

        record_id = self.tree.item(selected_item)["values"][0]

        if messagebox.askyesno("Xác nhận", "Bạn chắc muốn xóa?"):
            try:
                self.cursor.execute("DELETE FROM benh_nhan WHERE id=?", (record_id,))
                self.conn.commit()
                self.load_data()
                self.clear_fields()
                messagebox.showinfo("Thành công", "Đã xóa.")
            except Exception as e:
                messagebox.showerror("Lỗi", str(e))

    def search_data(self):
        keyword = self.search_entry.get()

        self.cursor.execute("""
            SELECT * FROM benh_nhan
            WHERE ho_ten LIKE ? OR ma_bn LIKE ?
        """, ('%' + keyword + '%', '%' + keyword + '%'))

        rows = self.cursor.fetchall()

        for item in self.tree.get_children():
            self.tree.delete(item)

        for row in rows:
            self.tree.insert("", tk.END, values=row)

    def clear_fields(self):
        for key, widget in self.entries.items():
            if key == "gioi_tinh":
                widget.set("Nam")
            else:
                widget.delete(0, tk.END)

    def on_select_item(self, event):
        selected = self.tree.focus()
        if not selected:
            return

        values = self.tree.item(selected)["values"]
        keys = ["ma_bn", "ho_ten", "gioi_tinh", "ngay_sinh", "so_dt",
                "dia_chi", "chan_doan", "ngay_nhap_vien", "ghi_chu"]

        for i, key in enumerate(keys):
            widget = self.entries[key]
            if key == "gioi_tinh":
                widget.set(values[i + 1])
            else:
                widget.delete(0, tk.END)
                widget.insert(0, values[i + 1])

    def export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Danh sách bệnh nhân"

            columns = ("ID", "Mã BN", "Họ tên", "Giới tính", "Ngày sinh", "SĐT",
                       "Địa chỉ", "Chẩn đoán", "Ngày nhập viện", "Ghi chú")
            ws.append(columns)

            for row_id in self.tree.get_children():
                row = self.tree.item(row_id)['values']
                ws.append(row)

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=10):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            wb.save("danh_sach_benh_nhan.xlsx")
            messagebox.showinfo("Thành công", "Xuất Excel thành công!")

        except Exception as e:
            messagebox.showerror("Lỗi", str(e))

if __name__ == "__main__":
    root = tk.Tk()
    center_window(root, 900, 650)  
    app = PatientManagerApp(root)
    root.mainloop()
 